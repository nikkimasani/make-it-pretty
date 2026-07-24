import csv
import io
from datetime import date, datetime, time, timedelta

from openpyxl import load_workbook


def process_table(
    content: bytes,
    filename: str,
    sheet_name: str | None = None,
) -> tuple[list[dict[str, str]], str, dict[str, object]]:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext in ("xlsx", "xls"):
        return _process_excel(content, sheet_name)

    text = _decode_content(content)
    encoding = _detect_encoding(content)
    if ext == "tsv":
        rows = _parse_csv(text, delimiter="\t")
    else:
        delimiter = _detect_delimiter(text)
        rows = _parse_csv(text, delimiter=delimiter)

    header_map, data_rows = _extract_headers(rows)
    column_names = header_map or [str(i) for i in range(len(rows[0]))] if rows else []

    result: list[dict[str, str]] = []
    for row in data_rows:
        entry: dict[str, str] = {}
        for i, col in enumerate(column_names):
            entry[col] = row[i].strip() if i < len(row) else ""
        result.append(entry)

    metadata: dict[str, object] = {
        "rows": len(data_rows),
        "columns": len(column_names),
        "column_names": column_names,
        "encoding": encoding,
        "format": ext or "csv",
    }

    return result, ext or "csv", metadata


def _detect_encoding(content: bytes) -> str:
    if content[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"
    try:
        content.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        try:
            content.decode("latin-1")
            return "latin-1"
        except UnicodeDecodeError:
            try:
                content.decode("windows-1252")
                return "windows-1252"
            except UnicodeDecodeError:
                return "utf-8"


def _decode_content(content: bytes) -> str:
    encoding = _detect_encoding(content)
    return content.decode(encoding)


def _detect_delimiter(text: str) -> str:
    sample = text[:5000]
    lines_sample = [ln for ln in sample.split("\n") if ln.strip()][:10]
    if not lines_sample:
        return ","

    tab_count = sum(ln.count("\t") for ln in lines_sample)
    comma_count = sum(ln.count(",") for ln in lines_sample)

    # If tabs exist and are at least as frequent as commas, prefer tabs
    if tab_count > 0 and tab_count >= comma_count:
        return "\t"

    # Check if commas divide lines into roughly equal columns
    if comma_count > 0:
        col_counts = [ln.count(",") + 1 for ln in lines_sample]
        if max(col_counts) - min(col_counts) <= 1 and max(col_counts) > 1:
            return ","
        # Could be TSV with no tabs in sample — check for \t in full text
        if "\t" in text:
            return "\t"

    return ","


def _parse_csv(text: str, delimiter: str = ",") -> list[list[str]]:
    # Filter empty trailing lines to avoid blank rows
    text = text.rstrip("\n")
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [row for row in reader if any(cell.strip() for cell in row)]


def _extract_headers(
    rows: list[list[str]],
) -> tuple[list[str] | None, list[list[str]]]:
    if not rows:
        return None, []

    first = rows[0]
    if all(
        v.strip()
        and len(v.strip()) < 100
        and not _looks_like_data(v.strip())
        for v in first
    ):
        unique = len(set(v.strip().lower() for v in first))
        if unique == len(first) and unique > 0:
            return [v.strip() for v in first], rows[1:]

    return None, rows


def _looks_like_data(value: str) -> bool:
    try:
        float(value.replace(",", "").replace("$", "").replace("€", "").replace("£", ""))
        return True
    except ValueError:
        pass
    if value.startswith(("http://", "https://", "ftp://")):
        return True
    return False


def _process_excel(
    content: bytes,
    sheet_name: str | None,
) -> tuple[list[dict[str, str]], str, dict[str, object]]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    sheet_names = wb.sheetnames

    if sheet_name and sheet_name in sheet_names:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    raw_rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        raw_rows.append([_format_cell(cell) for cell in row])

    # Remove trailing empty rows
    while raw_rows and all(c == "" for c in raw_rows[-1]):
        raw_rows.pop()
    # Remove trailing empty columns
    if raw_rows:
        max_cols = max(len(r) for r in raw_rows)
        for row in raw_rows:
            while len(row) < max_cols:
                row.append("")
        while max_cols > 0 and all(r[max_cols - 1] == "" for r in raw_rows):
            max_cols -= 1
            for row in raw_rows:
                row.pop()

    header_map, data_rows = _extract_headers(raw_rows)
    column_names = header_map or [str(i) for i in range(len(raw_rows[0]))] if raw_rows else []

    result: list[dict[str, str]] = []
    for row in data_rows:
        entry: dict[str, str] = {}
        for i, col in enumerate(column_names):
            entry[col] = row[i].strip() if i < len(row) else ""
        result.append(entry)

    active_idx = sheet_names.index(ws.title) if ws.title in sheet_names else 0

    metadata: dict[str, object] = {
        "rows": len(data_rows),
        "columns": len(column_names),
        "column_names": column_names,
        "encoding": "utf-8",
        "format": "xlsx",
        "sheet_names": sheet_names,
        "active_sheet": ws.title,
        "active_sheet_index": active_idx,
    }

    return result, "xlsx", metadata


def _format_cell(cell: object) -> str:
    if cell is None:
        return ""
    if isinstance(cell, datetime):
        if cell.hour == 0 and cell.minute == 0 and cell.second == 0 and cell.microsecond == 0:
            return cell.date().isoformat()
        return cell.isoformat()
    if isinstance(cell, date):
        return cell.isoformat()
    if isinstance(cell, time):
        return cell.isoformat()
    if isinstance(cell, timedelta):
        total = cell.total_seconds()
        hours = int(total // 3600)
        minutes = int((total % 3600) // 60)
        seconds = total % 60
        return f"{hours:02d}:{minutes:02d}:{seconds:05.2f}"
    return str(cell)
