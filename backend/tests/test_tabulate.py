import io
from datetime import date, datetime

from openpyxl import Workbook

from app.services.tabulate import (
    _detect_delimiter,
    _detect_encoding,
    _extract_headers,
    process_table,
)


def _make_excel(sheet_data: dict[str, list[tuple]], active_sheet: str | None = None) -> bytes:
    wb = Workbook()
    if active_sheet:
        wb.active = wb.sheetnames.index(active_sheet) if active_sheet in wb.sheetnames else 0
    for idx, (name, rows) in enumerate(sheet_data.items()):
        if idx == 0:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestCSV:
    def test_basic_csv(self):
        content = b"name,age,city\nAlice,30,NYC\nBob,25,LA\n"
        rows, fmt, meta = process_table(content, "data.csv")
        assert fmt == "csv"
        assert meta["rows"] == 2
        assert meta["columns"] == 3
        assert meta["column_names"] == ["name", "age", "city"]
        assert rows[0]["name"] == "Alice"
        assert rows[1]["city"] == "LA"

    def test_csv_with_bom(self):
        content = b"\xef\xbb\xbfname,age\nAlice,30\n"
        rows, fmt, meta = process_table(content, "data.csv")
        assert rows[0]["name"] == "Alice"
        assert meta["encoding"] == "utf-8-sig"

    def test_csv_quoted_fields(self):
        content = b'name,note\nAlice,"Hello, world"\nBob,"Line1\nLine2"\n'
        rows, fmt, meta = process_table(content, "data.csv")
        assert rows[0]["note"] == "Hello, world"
        assert rows[1]["note"] == "Line1\nLine2"

    def test_csv_no_headers_all_data(self):
        content = b"123,456,789\n321,654,987\n"
        rows, fmt, meta = process_table(content, "data.csv")
        # First row looks like data (numeric) -> no headers
        assert meta["column_names"] != ["123", "456", "789"]
        assert len(meta["column_names"]) == 3
        assert meta["rows"] == 2  # both rows are data

    def test_csv_single_column(self):
        content = b"value\n1\n2\n3\n"
        rows, fmt, meta = process_table(content, "data.csv")
        assert meta["columns"] == 1
        assert meta["rows"] == 3
        assert rows[0]["value"] == "1"

    def test_csv_empty(self):
        content = b""
        rows, fmt, meta = process_table(content, "")
        assert rows == []
        assert meta["rows"] == 0

    def test_csv_latin1_encoding(self):
        content = "name,description\nCafé,100%\n".encode("latin-1")
        rows, fmt, meta = process_table(content, "data.csv")
        assert meta["encoding"] == "latin-1"
        assert rows[0]["name"] == "Café"

    def test_csv_delimiter_auto_detect_tab(self):
        content = b"name\tage\tcity\nAlice\t30\tNYC\n"
        rows, fmt, meta = process_table(content, "data.csv")
        assert meta["rows"] == 1
        assert rows[0]["name"] == "Alice"


class TestTSV:
    def test_basic_tsv(self):
        content = b"name\tage\tcity\nAlice\t30\tNYC\nBob\t25\tLA\n"
        rows, fmt, meta = process_table(content, "data.tsv")
        assert fmt == "tsv"
        assert meta["rows"] == 2
        assert rows[0]["name"] == "Alice"

    def test_tsv_quoted(self):
        content = b'name\tnote\nAlice\t"Hello\tWorld"\n'
        rows, fmt, meta = process_table(content, "data.tsv")
        assert rows[0]["note"] == "Hello\tWorld"


class TestExcel:
    def test_basic_excel(self):
        xlsx = _make_excel({"Sheet1": [("name", "age"), ("Alice", 30), ("Bob", 25)]})
        rows, fmt, meta = process_table(xlsx, "data.xlsx")
        assert fmt == "xlsx"
        assert meta["rows"] == 2
        assert meta["column_names"] == ["name", "age"]
        assert rows[0]["name"] == "Alice"
        assert rows[1]["age"] == "25"

    def test_excel_multiple_sheets(self):
        xlsx = _make_excel({
            "Sheet1": [("col",), ("A",)],
            "Sheet2": [("val",), ("X",)],
        })
        rows, fmt, meta = process_table(xlsx, "data.xlsx")
        assert meta["sheet_names"] == ["Sheet1", "Sheet2"]
        assert meta["active_sheet"] == "Sheet1"
        # First sheet by default
        assert rows[0]["col"] == "A"

    def test_excel_select_sheet(self):
        xlsx = _make_excel({
            "Sheet1": [("col",), ("A",)],
            "Sheet2": [("val",), ("X",)],
        })
        rows, fmt, meta = process_table(xlsx, "data.xlsx", sheet_name="Sheet2")
        assert rows[0]["val"] == "X"

    def test_excel_merged_cells(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(("name", "value"))
        ws.append(("A", 1))
        ws.merge_cells("A2:A3")
        # MergedCell row 3 is populated by the merge; B goes to row 4
        ws.append(("B", 2))
        buf = io.BytesIO()
        wb.save(buf)
        xlsx = buf.getvalue()

        rows, fmt, meta = process_table(xlsx, "data.xlsx")
        # openpyxl data_only mode may skip fully-None merged rows
        assert meta["rows"] >= 2
        assert rows[0]["name"] == "A"

    def test_excel_date_time(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(("date", "datetime"))
        ws.append((date(2024, 1, 15), datetime(2024, 6, 1, 14, 30, 0)))
        buf = io.BytesIO()
        wb.save(buf)
        xlsx = buf.getvalue()

        rows, fmt, meta = process_table(xlsx, "data.xlsx")
        assert rows[0]["date"] == "2024-01-15"
        assert rows[0]["datetime"] == "2024-06-01T14:30:00"

    def test_excel_empty(self):
        content = _make_excel({"Sheet1": []})
        rows, fmt, meta = process_table(content, "data.xlsx")
        assert rows == []
        assert meta["rows"] == 0


class TestDetection:
    def test_detect_encoding_utf8(self):
        assert _detect_encoding(b"hello") == "utf-8"

    def test_detect_encoding_latin1(self):
        assert _detect_encoding("Café".encode("latin-1")) == "latin-1"

    def test_detect_delimiter_comma(self):
        assert _detect_delimiter("a,b,c\n1,2,3\n") == ","

    def test_detect_delimiter_tab(self):
        assert _detect_delimiter("a\tb\tc\n1\t2\t3\n") == "\t"

    def test_extract_headers_with_data_looking_values(self):
        rows = [["123", "456"], ["789", "012"]]
        headers, data = _extract_headers(rows)
        assert headers is None  # numeric headers = no headers
        assert data == rows

    def test_extract_headers_valid(self):
        rows = [["Name", "Age"], ["Alice", "30"]]
        headers, data = _extract_headers(rows)
        assert headers == ["Name", "Age"]
        assert data == [["Alice", "30"]]
